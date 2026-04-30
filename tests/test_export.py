"""Tests for data export module — see SPECS.md §35."""

import json
from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest

from tallybridge.cache import TallyCache
from tallybridge.export import DataExporter
from tallybridge.models.master import TallyLedger
from tallybridge.models.voucher import (
    TallyVoucher,
    TallyVoucherEntry,
)


@pytest.fixture
def populated_cache(tmp_path: Path) -> TallyCache:
    db = TallyCache(str(tmp_path / "test.duckdb"))
    db.upsert_ledgers(
        [
            TallyLedger(
                name="Cash",
                guid="g-cash",
                alter_id=1,
                parent_group="Cash-in-Hand",
                closing_balance=Decimal("50000"),
            ),
            TallyLedger(
                name="Sales",
                guid="g-sales",
                alter_id=2,
                parent_group="Sales Accounts",
                closing_balance=Decimal("120000"),
            ),
            TallyLedger(
                name="Rent Expense",
                guid="g-rent",
                alter_id=3,
                parent_group="Indirect Expenses",
                closing_balance=Decimal("10000"),
            ),
        ]
    )
    db.upsert_vouchers(
        [
            TallyVoucher(
                guid="v-001",
                alter_id=100,
                voucher_number="SI/001",
                voucher_type="Sales",
                date=date(2026, 4, 1),
                party_ledger="Customer A",
                total_amount=Decimal("59000"),
                ledger_entries=[
                    TallyVoucherEntry(
                        ledger_name="Customer A", amount=Decimal("59000")
                    ),
                    TallyVoucherEntry(
                        ledger_name="Sales", amount=Decimal("-50000")
                    ),
                ],
            ),
        ]
    )
    return db


@pytest.fixture
def exporter(populated_cache: TallyCache) -> DataExporter:
    return DataExporter(populated_cache)


def test_export_csv_ledgers(exporter: DataExporter, tmp_path: Path) -> None:
    output = tmp_path / "ledgers.csv"
    count = exporter.export_csv("ledgers", str(output))
    assert count == 3
    content = output.read_text(encoding="utf-8")
    assert "Cash" in content
    assert "Sales" in content
    assert "guid" in content


def test_export_csv_with_filter(exporter: DataExporter, tmp_path: Path) -> None:
    output = tmp_path / "filtered.csv"
    count = exporter.export_csv(
        "ledgers", str(output), where="name = 'Cash'"
    )
    assert count == 1
    content = output.read_text(encoding="utf-8")
    assert "Cash" in content
    assert "Sales" not in content


def test_export_csv_with_columns(exporter: DataExporter, tmp_path: Path) -> None:
    output = tmp_path / "cols.csv"
    count = exporter.export_csv(
        "ledgers", str(output), columns=["name", "guid"]
    )
    assert count == 3
    content = output.read_text(encoding="utf-8")
    lines = content.strip().split("\n")
    assert "name" in lines[0]
    assert "guid" in lines[0]
    assert "closing_balance" not in lines[0]


def test_export_csv_with_limit(exporter: DataExporter, tmp_path: Path) -> None:
    output = tmp_path / "limited.csv"
    count = exporter.export_csv("ledgers", str(output), limit=2)
    assert count == 2


def test_export_csv_invalid_table(exporter: DataExporter, tmp_path: Path) -> None:
    with pytest.raises(ValueError, match="Unknown table"):
        exporter.export_csv("nonexistent", str(tmp_path / "bad.csv"))


def test_export_json_vouchers(exporter: DataExporter, tmp_path: Path) -> None:
    output = tmp_path / "vouchers.json"
    count = exporter.export_json("vouchers", str(output))
    assert count == 1
    data = json.loads(output.read_text(encoding="utf-8"))
    assert isinstance(data, list)
    assert len(data) == 1
    assert data[0]["voucher_type"] == "Sales"


def test_export_json_with_filter(exporter: DataExporter, tmp_path: Path) -> None:
    output = tmp_path / "filtered.json"
    count = exporter.export_json(
        "ledgers", str(output), where="name = 'Sales'"
    )
    assert count == 1
    data = json.loads(output.read_text(encoding="utf-8"))
    assert data[0]["name"] == "Sales"


def test_export_empty_table(exporter: DataExporter, tmp_path: Path) -> None:
    output = tmp_path / "empty.csv"
    count = exporter.export_csv("godowns", str(output))
    assert count == 0

    json_output = tmp_path / "empty.json"
    count = exporter.export_json("godowns", str(json_output))
    assert count == 0
    data = json.loads(json_output.read_text(encoding="utf-8"))
    assert data == []


def test_export_excel_multi_table(exporter: DataExporter, tmp_path: Path) -> None:
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        pytest.skip("openpyxl not installed")

    output = tmp_path / "multi.xlsx"
    result = exporter.export_excel(str(output), tables=["ledgers", "vouchers"])
    assert result["ledgers"] == 3
    assert result["vouchers"] == 1

    import openpyxl

    wb = openpyxl.load_workbook(str(output))
    assert "ledgers" in wb.sheetnames
    assert "vouchers" in wb.sheetnames


def test_export_excel_default_tables(exporter: DataExporter, tmp_path: Path) -> None:
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        pytest.skip("openpyxl not installed")

    output = tmp_path / "default.xlsx"
    result = exporter.export_excel(str(output))
    assert len(result) > 0


def test_export_excel_missing_openpyxl(
    exporter: DataExporter, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    import sys

    openpyxl_mod = sys.modules.get("openpyxl")
    sys.modules["openpyxl"] = None  # type: ignore[assignment]
    try:
        with pytest.raises(ImportError, match="openpyxl"):
            exporter.export_excel(str(tmp_path / "fail.xlsx"))
    finally:
        if openpyxl_mod is not None:
            sys.modules["openpyxl"] = openpyxl_mod
        else:
            del sys.modules["openpyxl"]


def test_export_csv_bytes(exporter: DataExporter) -> None:
    result = exporter.export_csv_bytes("ledgers")
    assert "Cash" in result
    assert "Sales" in result
    assert "\n" in result


def test_export_csv_bytes_empty(exporter: DataExporter) -> None:
    result = exporter.export_csv_bytes("godowns")
    assert result == ""


def test_export_creates_parent_dirs(exporter: DataExporter, tmp_path: Path) -> None:
    output = tmp_path / "subdir" / "deep" / "ledgers.csv"
    count = exporter.export_csv("ledgers", str(output))
    assert count == 3
    assert output.exists()


def test_export_csv_chunked(exporter: DataExporter, tmp_path: Path) -> None:
    output = tmp_path / "chunked.csv"
    count = exporter.export_csv_chunked(
        "ledgers", str(output), chunk_size=2
    )
    assert count == 3
    content = output.read_text(encoding="utf-8")
    assert "Cash" in content
    assert "Sales" in content
